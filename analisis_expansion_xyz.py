#!/usr/bin/env python3
"""
Análisis de Expansión Regional - Corporación XYZ
Análisis de decisiones bajo incertidumbre y riesgo para selección de ciudad en la costa Caribe
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference

# Datos del problema
INVERSION_INICIAL = 800000  # USD

# Costos operativos mensuales (en pesos colombianos)
costos_operativos = {
    'Santa Marta': {'arriendo': 2000000, 'servicios': 3500000, 'mantenimiento': 4000000},
    'Barranquilla': {'arriendo': 2200000, 'servicios': 3000000, 'mantenimiento': 3800000},
    'Cartagena': {'arriendo': 2800000, 'servicios': 4000000, 'mantenimiento': 3500000}
}

# Ingresos proyectados mensuales (en pesos colombianos)
ingresos_proyectados = {
    'Santa Marta': {'alta': 15000000, 'media': 12000000, 'baja': 5000000},
    'Barranquilla': {'alta': 14000000, 'media': 10000000, 'baja': 5000000},
    'Cartagena': {'alta': 16000000, 'media': 8000000, 'baja': 6000000}
}

# Probabilidades iniciales y nuevas
prob_iniciales = {'alta': 0.2, 'media': 0.3, 'baja': 0.5}
prob_nuevas = {'alta': 0.3, 'media': 0.4, 'baja': 0.3}

# Costo del estudio técnico
costo_estudio = 1000000  # pesos

def crear_workbook():
    """Crear workbook de Excel con análisis completo"""
    wb = openpyxl.Workbook()
    
    # Eliminar hoja por defecto
    wb.remove(wb.active)
    
    # Crear hojas
    ws_datos = wb.create_sheet("1. Datos Base")
    ws_criterios = wb.create_sheet("2. Criterios Decision")
    ws_arbol = wb.create_sheet("3. Arbol Decisiones")
    ws_estudio = wb.create_sheet("4. Con Estudio Tecnico")
    ws_recomendaciones = wb.create_sheet("5. Recomendaciones")
    
    # Configurar cada hoja
    configurar_datos_base(ws_datos)
    configurar_criterios_decision(ws_criterios)
    configurar_arbol_decisiones(ws_arbol)
    configurar_con_estudio(ws_estudio)
    configurar_recomendaciones(ws_recomendaciones)
    
    return wb

def aplicar_estilo_header(ws, cell, color="366092"):
    """Aplicar estilo de encabezado"""
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

def configurar_datos_base(ws):
    """Configurar hoja de datos base"""
    # Título
    ws['A1'] = "ANÁLISIS DE EXPANSIÓN REGIONAL - CORPORACIÓN XYZ"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Información general
    ws['A3'] = "INFORMACIÓN GENERAL"
    aplicar_estilo_header(ws, ws['A3'])
    ws.merge_cells('A3:B3')
    
    ws['A4'] = "Inversión Inicial (USD):"
    ws['B4'] = INVERSION_INICIAL
    ws['B4'].number_format = '$#,##0'
    
    # Costos operativos mensuales
    ws['A6'] = "COSTOS OPERATIVOS MENSUALES (COP)"
    aplicar_estilo_header(ws, ws['A6'])
    ws.merge_cells('A6:E6')
    
    headers = ['Ciudad', 'Arriendo', 'Servicios Públicos', 'Mantenimiento', 'Total']
    for i, header in enumerate(headers):
        cell = ws.cell(row=7, column=i+1, value=header)
        aplicar_estilo_header(ws, cell, "4472C4")
    
    row = 8
    for ciudad, costos in costos_operativos.items():
        ws.cell(row=row, column=1, value=ciudad)
        ws.cell(row=row, column=2, value=costos['arriendo']).number_format = '$#,##0'
        ws.cell(row=row, column=3, value=costos['servicios']).number_format = '$#,##0'
        ws.cell(row=row, column=4, value=costos['mantenimiento']).number_format = '$#,##0'
        total = costos['arriendo'] + costos['servicios'] + costos['mantenimiento']
        ws.cell(row=row, column=5, value=total).number_format = '$#,##0'
        row += 1
    
    # Ingresos proyectados
    ws['A12'] = "INGRESOS PROYECTADOS MENSUALES (COP)"
    aplicar_estilo_header(ws, ws['A12'])
    ws.merge_cells('A12:E12')
    
    headers = ['Ciudad', 'Demanda Alta', 'Demanda Media', 'Demanda Baja']
    for i, header in enumerate(headers):
        cell = ws.cell(row=13, column=i+1, value=header)
        aplicar_estilo_header(ws, cell, "4472C4")
    
    row = 14
    for ciudad, ingresos in ingresos_proyectados.items():
        ws.cell(row=row, column=1, value=ciudad)
        ws.cell(row=row, column=2, value=ingresos['alta']).number_format = '$#,##0'
        ws.cell(row=row, column=3, value=ingresos['media']).number_format = '$#,##0'
        ws.cell(row=row, column=4, value=ingresos['baja']).number_format = '$#,##0'
        row += 1
    
    # Probabilidades
    ws['A18'] = "PROBABILIDADES DE ESCENARIOS"
    aplicar_estilo_header(ws, ws['A18'])
    ws.merge_cells('A18:D18')
    
    ws['A19'] = "Escenario"
    ws['B19'] = "Probabilidad Inicial"
    ws['C19'] = "Probabilidad Nueva"
    aplicar_estilo_header(ws, ws['A19'], "4472C4")
    aplicar_estilo_header(ws, ws['B19'], "4472C4")
    aplicar_estilo_header(ws, ws['C19'], "4472C4")
    
    escenarios = ['Demanda Alta', 'Demanda Media', 'Demanda Baja']
    prob_init = [0.2, 0.3, 0.5]
    prob_new = [0.3, 0.4, 0.3]
    
    for i, (escenario, p_init, p_new) in enumerate(zip(escenarios, prob_init, prob_new)):
        row = 20 + i
        ws.cell(row=row, column=1, value=escenario)
        ws.cell(row=row, column=2, value=p_init).number_format = '0.0%'
        ws.cell(row=row, column=3, value=p_new).number_format = '0.0%'

def calcular_matriz_pagos():
    """Calcular matriz de pagos (utilidad neta mensual)"""
    matriz = {}
    
    for ciudad in costos_operativos.keys():
        matriz[ciudad] = {}
        costo_total = sum(costos_operativos[ciudad].values())
        
        for escenario in ['alta', 'media', 'baja']:
            ingreso = ingresos_proyectados[ciudad][escenario]
            utilidad = ingreso - costo_total
            matriz[ciudad][escenario] = utilidad
    
    return matriz

def configurar_criterios_decision(ws):
    """Configurar hoja de criterios de decisión"""
    # Título
    ws['A1'] = "CRITERIOS DE DECISIÓN BAJO INCERTIDUMBRE Y RIESGO"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Matriz de pagos
    ws['A3'] = "MATRIZ DE PAGOS (Utilidad Neta Mensual - COP)"
    aplicar_estilo_header(ws, ws['A3'])
    ws.merge_cells('A3:E3')
    
    matriz = calcular_matriz_pagos()
    
    # Headers de la matriz
    headers = ['Ciudad', 'Demanda Alta', 'Demanda Media', 'Demanda Baja']
    for i, header in enumerate(headers):
        cell = ws.cell(row=4, column=i+1, value=header)
        aplicar_estilo_header(ws, cell, "4472C4")
    
    # Llenar matriz
    row = 5
    for ciudad in matriz.keys():
        ws.cell(row=row, column=1, value=ciudad)
        ws.cell(row=row, column=2, value=matriz[ciudad]['alta']).number_format = '$#,##0'
        ws.cell(row=row, column=3, value=matriz[ciudad]['media']).number_format = '$#,##0'
        ws.cell(row=row, column=4, value=matriz[ciudad]['baja']).number_format = '$#,##0'
        row += 1
    
    # Criterios de decisión
    row += 2
    ws.cell(row=row, column=1, value="CRITERIOS DE DECISIÓN")
    aplicar_estilo_header(ws, ws.cell(row=row, column=1))
    ws.merge_cells(f'A{row}:F{row}')
    
    # Preparar datos para cálculos
    ciudades = list(matriz.keys())
    valores = {ciudad: list(matriz[ciudad].values()) for ciudad in ciudades}
    
    # 1. Criterio de Laplace (Equiprobabilidad)
    row += 2
    ws.cell(row=row, column=1, value="1. CRITERIO DE LAPLACE (Equiprobabilidad)")
    aplicar_estilo_header(ws, ws.cell(row=row, column=1), "70AD47")
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    ws.cell(row=row, column=1, value="Ciudad")
    ws.cell(row=row, column=2, value="Valor Esperado")
    aplicar_estilo_header(ws, ws.cell(row=row, column=1), "93C47D")
    aplicar_estilo_header(ws, ws.cell(row=row, column=2), "93C47D")
    
    mejor_laplace = ""
    max_laplace = float('-inf')
    
    for ciudad in ciudades:
        row += 1
        promedio = sum(valores[ciudad]) / len(valores[ciudad])
        ws.cell(row=row, column=1, value=ciudad)
        ws.cell(row=row, column=2, value=promedio).number_format = '$#,##0'
        
        if promedio > max_laplace:
            max_laplace = promedio
            mejor_laplace = ciudad
    
    row += 1
    ws.cell(row=row, column=1, value="MEJOR ALTERNATIVA:")
    ws.cell(row=row, column=2, value=mejor_laplace)
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).font = Font(bold=True, color="70AD47")
    
    # 2. Criterio Maximax (Optimista)
    row += 2
    ws.cell(row=row, column=1, value="2. CRITERIO MAXIMAX (Optimista)")
    aplicar_estilo_header(ws, ws.cell(row=row, column=1), "FFC000")
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    ws.cell(row=row, column=1, value="Ciudad")
    ws.cell(row=row, column=2, value="Mejor Resultado")
    aplicar_estilo_header(ws, ws.cell(row=row, column=1), "FFD966")
    aplicar_estilo_header(ws, ws.cell(row=row, column=2), "FFD966")
    
    mejor_maximax = ""
    max_maximax = float('-inf')
    
    for ciudad in ciudades:
        row += 1
        maximo = max(valores[ciudad])
        ws.cell(row=row, column=1, value=ciudad)
        ws.cell(row=row, column=2, value=maximo).number_format = '$#,##0'
        
        if maximo > max_maximax:
            max_maximax = maximo
            mejor_maximax = ciudad
    
    row += 1
    ws.cell(row=row, column=1, value="MEJOR ALTERNATIVA:")
    ws.cell(row=row, column=2, value=mejor_maximax)
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).font = Font(bold=True, color="FFC000")
    
    # 3. Criterio Maximin (Pesimista)
    row += 2
    ws.cell(row=row, column=1, value="3. CRITERIO MAXIMIN (Pesimista)")
    aplicar_estilo_header(ws, ws.cell(row=row, column=1), "C55A5A")
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    ws.cell(row=row, column=1, value="Ciudad")
    ws.cell(row=row, column=2, value="Peor Resultado")
    aplicar_estilo_header(ws, ws.cell(row=row, column=1), "E06666")
    aplicar_estilo_header(ws, ws.cell(row=row, column=2), "E06666")
    
    mejor_maximin = ""
    max_maximin = float('-inf')
    
    for ciudad in ciudades:
        row += 1
        minimo = min(valores[ciudad])
        ws.cell(row=row, column=1, value=ciudad)
        ws.cell(row=row, column=2, value=minimo).number_format = '$#,##0'
        
        if minimo > max_maximin:
            max_maximin = minimo
            mejor_maximin = ciudad
    
    row += 1
    ws.cell(row=row, column=1, value="MEJOR ALTERNATIVA:")
    ws.cell(row=row, column=2, value=mejor_maximin)
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).font = Font(bold=True, color="C55A5A")
    
    # 4. Criterio de Hurwicz
    row += 2
    alpha = 0.6  # Coeficiente de optimismo
    ws.cell(row=row, column=1, value=f"4. CRITERIO DE HURWICZ (α = {alpha})")
    aplicar_estilo_header(ws, ws.cell(row=row, column=1), "9900FF")
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    ws.cell(row=row, column=1, value="Ciudad")
    ws.cell(row=row, column=2, value="Valor Hurwicz")
    aplicar_estilo_header(ws, ws.cell(row=row, column=1), "B266CC")
    aplicar_estilo_header(ws, ws.cell(row=row, column=2), "B266CC")
    
    mejor_hurwicz = ""
    max_hurwicz = float('-inf')
    
    for ciudad in ciudades:
        row += 1
        maximo = max(valores[ciudad])
        minimo = min(valores[ciudad])
        hurwicz = alpha * maximo + (1 - alpha) * minimo
        ws.cell(row=row, column=1, value=ciudad)
        ws.cell(row=row, column=2, value=hurwicz).number_format = '$#,##0'
        
        if hurwicz > max_hurwicz:
            max_hurwicz = hurwicz
            mejor_hurwicz = ciudad
    
    row += 1
    ws.cell(row=row, column=1, value="MEJOR ALTERNATIVA:")
    ws.cell(row=row, column=2, value=mejor_hurwicz)
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).font = Font(bold=True, color="9900FF")
    
    # 5. Criterio de Savage (Minimax Arrepentimiento)
    row += 2
    ws.cell(row=row, column=1, value="5. CRITERIO DE SAVAGE (Minimax Arrepentimiento)")
    aplicar_estilo_header(ws, ws.cell(row=row, column=1), "FF6D01")
    ws.merge_cells(f'A{row}:F{row}')
    
    # Calcular matriz de arrepentimiento
    max_por_escenario = []
    escenarios = ['alta', 'media', 'baja']
    for i, escenario in enumerate(escenarios):
        max_val = max(matriz[ciudad][escenario] for ciudad in ciudades)
        max_por_escenario.append(max_val)
    
    # Mostrar matriz de arrepentimiento
    row += 1
    headers = ['Ciudad', 'Demanda Alta', 'Demanda Media', 'Demanda Baja', 'Máx Arrepent.']
    for i, header in enumerate(headers):
        cell = ws.cell(row=row, column=i+1, value=header)
        aplicar_estilo_header(ws, cell, "FF9500")
    
    mejor_savage = ""
    min_savage = float('inf')
    
    for ciudad in ciudades:
        row += 1
        ws.cell(row=row, column=1, value=ciudad)
        arrepentimientos = []
        
        for i, escenario in enumerate(escenarios):
            arrepentimiento = max_por_escenario[i] - matriz[ciudad][escenario]
            arrepentimientos.append(arrepentimiento)
            ws.cell(row=row, column=i+2, value=arrepentimiento).number_format = '$#,##0'
        
        max_arrepentimiento = max(arrepentimientos)
        ws.cell(row=row, column=5, value=max_arrepentimiento).number_format = '$#,##0'
        
        if max_arrepentimiento < min_savage:
            min_savage = max_arrepentimiento
            mejor_savage = ciudad
    
    row += 1
    ws.cell(row=row, column=1, value="MEJOR ALTERNATIVA:")
    ws.cell(row=row, column=2, value=mejor_savage)
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).font = Font(bold=True, color="FF6D01")
    
    # 6. Criterio del Valor Esperado
    row += 2
    ws.cell(row=row, column=1, value="6. CRITERIO DEL VALOR ESPERADO")
    aplicar_estilo_header(ws, ws.cell(row=row, column=1), "00B0F0")
    ws.merge_cells(f'A{row}:D{row}')
    
    row += 1
    ws.cell(row=row, column=1, value="Ciudad")
    ws.cell(row=row, column=2, value="Valor Esperado")
    aplicar_estilo_header(ws, ws.cell(row=row, column=1), "4DC2E8")
    aplicar_estilo_header(ws, ws.cell(row=row, column=2), "4DC2E8")
    
    mejor_ve = ""
    max_ve = float('-inf')
    
    prob_list = [prob_iniciales['alta'], prob_iniciales['media'], prob_iniciales['baja']]
    
    for ciudad in ciudades:
        row += 1
        valor_esperado = sum(prob_list[i] * valores[ciudad][i] for i in range(3))
        ws.cell(row=row, column=1, value=ciudad)
        ws.cell(row=row, column=2, value=valor_esperado).number_format = '$#,##0'
        
        if valor_esperado > max_ve:
            max_ve = valor_esperado
            mejor_ve = ciudad
    
    row += 1
    ws.cell(row=row, column=1, value="MEJOR ALTERNATIVA:")
    ws.cell(row=row, column=2, value=mejor_ve)
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).font = Font(bold=True, color="00B0F0")
    
    # Resumen de criterios
    row += 3
    ws.cell(row=row, column=1, value="RESUMEN DE CRITERIOS DE DECISIÓN")
    aplicar_estilo_header(ws, ws.cell(row=row, column=1), "366092")
    ws.merge_cells(f'A{row}:C{row}')
    
    criterios_resultados = [
        ("Laplace", mejor_laplace),
        ("Maximax", mejor_maximax),
        ("Maximin", mejor_maximin),
        ("Hurwicz", mejor_hurwicz),
        ("Savage", mejor_savage),
        ("Valor Esperado", mejor_ve)
    ]
    
    row += 1
    ws.cell(row=row, column=1, value="Criterio")
    ws.cell(row=row, column=2, value="Mejor Alternativa")
    aplicar_estilo_header(ws, ws.cell(row=row, column=1), "4472C4")
    aplicar_estilo_header(ws, ws.cell(row=row, column=2), "4472C4")
    
    for criterio, resultado in criterios_resultados:
        row += 1
        ws.cell(row=row, column=1, value=criterio)
        ws.cell(row=row, column=2, value=resultado)

def configurar_arbol_decisiones(ws):
    """Configurar hoja de árbol de decisiones"""
    ws['A1'] = "ANÁLISIS DEL ÁRBOL DE DECISIONES"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Análisis sin estudio técnico
    ws['A3'] = "DECISIÓN SIN ESTUDIO TÉCNICO"
    aplicar_estilo_header(ws, ws['A3'], "366092")
    ws.merge_cells('A3:H3')
    
    matriz = calcular_matriz_pagos()
    
    # Cálculo del valor esperado para cada ciudad
    row = 5
    ws.cell(row=row, column=1, value="Ciudad")
    ws.cell(row=row, column=2, value="VE Demanda Alta")
    ws.cell(row=row, column=3, value="VE Demanda Media")
    ws.cell(row=row, column=4, value="VE Demanda Baja")
    ws.cell(row=row, column=5, value="Valor Esperado Total")
    
    for i in range(5):
        aplicar_estilo_header(ws, ws.cell(row=row, column=i+1), "4472C4")
    
    mejor_ciudad_sin_estudio = ""
    max_ve_sin_estudio = float('-inf')
    
    prob_list = [prob_iniciales['alta'], prob_iniciales['media'], prob_iniciales['baja']]
    escenarios = ['alta', 'media', 'baja']
    
    for ciudad in matriz.keys():
        row += 1
        ws.cell(row=row, column=1, value=ciudad)
        
        ve_total = 0
        for i, escenario in enumerate(escenarios):
            ve_escenario = prob_list[i] * matriz[ciudad][escenario]
            ws.cell(row=row, column=i+2, value=ve_escenario).number_format = '$#,##0'
            ve_total += ve_escenario
        
        ws.cell(row=row, column=5, value=ve_total).number_format = '$#,##0'
        
        if ve_total > max_ve_sin_estudio:
            max_ve_sin_estudio = ve_total
            mejor_ciudad_sin_estudio = ciudad
    
    row += 2
    ws.cell(row=row, column=1, value="MEJOR DECISIÓN SIN ESTUDIO:")
    ws.cell(row=row, column=2, value=mejor_ciudad_sin_estudio)
    ws.cell(row=row, column=3, value=max_ve_sin_estudio).number_format = '$#,##0'
    
    for i in range(3):
        ws.cell(row=row, column=i+1).font = Font(bold=True)
    ws.cell(row=row, column=2).font = Font(bold=True, color="366092")
    
    # Representación gráfica del árbol
    row += 3
    ws.cell(row=row, column=1, value="ESTRUCTURA DEL ÁRBOL DE DECISIONES")
    aplicar_estilo_header(ws, ws.cell(row=row, column=1), "70AD47")
    ws.merge_cells(f'A{row}:H{row}')
    
    row += 2
    estructura_arbol = [
        "DECISIÓN INICIAL",
        "├── Santa Marta",
        "│   ├── Demanda Alta (20%) → Utilidad: $5,500,000",
        "│   ├── Demanda Media (30%) → Utilidad: $2,500,000", 
        "│   └── Demanda Baja (50%) → Utilidad: $-4,500,000",
        "├── Barranquilla", 
        "│   ├── Demanda Alta (20%) → Utilidad: $5,000,000",
        "│   ├── Demanda Media (30%) → Utilidad: $1,000,000",
        "│   └── Demanda Baja (50%) → Utilidad: $-4,000,000",
        "└── Cartagena",
        "    ├── Demanda Alta (20%) → Utilidad: $5,700,000", 
        "    ├── Demanda Media (30%) → Utilidad: $-2,300,000",
        "    └── Demanda Baja (50%) → Utilidad: $-4,300,000"
    ]
    
    for linea in estructura_arbol:
        ws.cell(row=row, column=1, value=linea)
        if "DECISIÓN" in linea:
            ws.cell(row=row, column=1).font = Font(bold=True, color="70AD47")
        elif linea.startswith("├──") or linea.startswith("└──"):
            ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1

def configurar_con_estudio(ws):
    """Configurar hoja con estudio técnico"""
    ws['A1'] = "ANÁLISIS CON ESTUDIO TÉCNICO"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Información del estudio
    ws['A3'] = f"COSTO DEL ESTUDIO TÉCNICO: ${costo_estudio:,}"
    ws['A3'].font = Font(bold=True, color="C55A5A")
    
    ws['A4'] = "NUEVAS PROBABILIDADES:"
    ws['B4'] = "Alta: 30%"
    ws['C4'] = "Media: 40%" 
    ws['D4'] = "Baja: 30%"
    ws['A4'].font = Font(bold=True)
    
    matriz = calcular_matriz_pagos()
    
    # Cálculo con nuevas probabilidades
    row = 7
    ws.cell(row=row, column=1, value="DECISIÓN CON NUEVAS PROBABILIDADES")
    aplicar_estilo_header(ws, ws.cell(row=row, column=1), "366092")
    ws.merge_cells(f'A{row}:F{row}')
    
    row += 1
    headers = ["Ciudad", "VE Demanda Alta", "VE Demanda Media", "VE Demanda Baja", "VE Total", "VE Neto*"]
    for i, header in enumerate(headers):
        aplicar_estilo_header(ws, ws.cell(row=row, column=i+1, value=header), "4472C4")
    
    mejor_ciudad_con_estudio = ""
    max_ve_con_estudio = float('-inf')
    max_ve_neto = float('-inf')
    
    prob_nuevas_list = [prob_nuevas['alta'], prob_nuevas['media'], prob_nuevas['baja']]
    escenarios = ['alta', 'media', 'baja']
    
    for ciudad in matriz.keys():
        row += 1
        ws.cell(row=row, column=1, value=ciudad)
        
        ve_total = 0
        for i, escenario in enumerate(escenarios):
            ve_escenario = prob_nuevas_list[i] * matriz[ciudad][escenario]
            ws.cell(row=row, column=i+2, value=ve_escenario).number_format = '$#,##0'
            ve_total += ve_escenario
        
        ws.cell(row=row, column=5, value=ve_total).number_format = '$#,##0'
        
        # VE neto (descontando costo del estudio)
        ve_neto = ve_total - costo_estudio
        ws.cell(row=row, column=6, value=ve_neto).number_format = '$#,##0'
        
        if ve_neto > max_ve_neto:
            max_ve_neto = ve_neto
            mejor_ciudad_con_estudio = ciudad
        
        if ve_total > max_ve_con_estudio:
            max_ve_con_estudio = ve_total
    
    row += 1
    ws.cell(row=row, column=1, value="*VE Neto = VE Total - Costo Estudio")
    ws.cell(row=row, column=1).font = Font(italic=True, size=9)
    
    row += 1
    ws.cell(row=row, column=1, value="MEJOR DECISIÓN CON ESTUDIO:")
    ws.cell(row=row, column=2, value=mejor_ciudad_con_estudio)
    ws.cell(row=row, column=3, value=max_ve_neto).number_format = '$#,##0'
    
    for i in range(3):
        ws.cell(row=row, column=i+1).font = Font(bold=True)
    ws.cell(row=row, column=2).font = Font(bold=True, color="366092")
    
    # Análisis comparativo
    row += 3
    ws.cell(row=row, column=1, value="ANÁLISIS COMPARATIVO")
    aplicar_estilo_header(ws, ws.cell(row=row, column=1), "70AD47")
    ws.merge_cells(f'A{row}:D{row}')
    
    # Calcular VE sin estudio para la mejor opción con nuevas probabilidades
    mejor_sin_estudio_ciudad = ""
    max_ve_sin_estudio_nuevas_prob = float('-inf')
    
    for ciudad in matriz.keys():
        ve_total = sum(prob_nuevas_list[i] * matriz[ciudad][escenarios[i]] for i in range(3))
        if ve_total > max_ve_sin_estudio_nuevas_prob:
            max_ve_sin_estudio_nuevas_prob = ve_total
            mejor_sin_estudio_ciudad = ciudad
    
    row += 2
    comparaciones = [
        ("Mejor opción SIN estudio (prob. nuevas):", mejor_sin_estudio_ciudad, max_ve_sin_estudio_nuevas_prob),
        ("Mejor opción CON estudio (neto):", mejor_ciudad_con_estudio, max_ve_neto),
        ("Valor de la información perfecta:", "Diferencia", max_ve_sin_estudio_nuevas_prob - max_ve_neto)
    ]
    
    for desc, ciudad, valor in comparaciones:
        ws.cell(row=row, column=1, value=desc)
        ws.cell(row=row, column=2, value=ciudad)
        ws.cell(row=row, column=3, value=valor).number_format = '$#,##0'
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
    
    # Recomendación
    row += 1
    if max_ve_neto > max_ve_sin_estudio_nuevas_prob:
        recomendacion = "SE RECOMIENDA REALIZAR EL ESTUDIO TÉCNICO"
        color = "70AD47"
    else:
        recomendacion = "NO SE RECOMIENDA REALIZAR EL ESTUDIO TÉCNICO"
        color = "C55A5A"
    
    ws.cell(row=row, column=1, value="RECOMENDACIÓN:")
    ws.cell(row=row, column=2, value=recomendacion)
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).font = Font(bold=True, color=color)

def configurar_recomendaciones(ws):
    """Configurar hoja de recomendaciones y conclusiones"""
    ws['A1'] = "RECOMENDACIONES Y CONCLUSIONES"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Recomendación principal
    row = 3
    ws.cell(row=row, column=1, value="RECOMENDACIÓN PRINCIPAL")
    aplicar_estilo_header(ws, ws.cell(row=row, column=1), "366092")
    ws.merge_cells(f'A{row}:G{row}')
    
    # Determinar la mejor opción basada en análisis
    matriz = calcular_matriz_pagos()
    prob_list = [prob_iniciales['alta'], prob_iniciales['media'], prob_iniciales['baja']]
    
    mejor_ciudad = ""
    max_ve = float('-inf')
    
    for ciudad in matriz.keys():
        ve = sum(prob_list[i] * list(matriz[ciudad].values())[i] for i in range(3))
        if ve > max_ve:
            max_ve = ve
            mejor_ciudad = ciudad
    
    recomendaciones = [
        "",
        f"Basado en el análisis integral de criterios de decisión, se RECOMIENDA seleccionar {mejor_ciudad.upper()} como ubicación para la primera tienda regional de Corporación XYZ en la costa Caribe.",
        "",
        "JUSTIFICACIÓN:",
        f"• {mejor_ciudad} presenta el mayor valor esperado (${max_ve:,.0f}) considerando las probabilidades iniciales de demanda",
        "• Esta ciudad ofrece el mejor balance entre potencial de ingresos y control de costos operativos",
        "• La decisión es consistente con múltiples criterios de evaluación analizados"
    ]
    
    for rec in recomendaciones:
        row += 1
        ws.cell(row=row, column=1, value=rec)
        if "RECOMIENDA" in rec:
            ws.cell(row=row, column=1).font = Font(bold=True, color="366092")
            ws.merge_cells(f'A{row}:G{row}')
        elif "JUSTIFICACIÓN" in rec:
            ws.cell(row=row, column=1).font = Font(bold=True)
        elif rec.startswith("•"):
            ws.merge_cells(f'A{row}:G{row}')
    
    # Fortalezas del método
    row += 2
    ws.cell(row=row, column=1, value="FORTALEZAS DEL MÉTODO DE ÁRBOLES DE DECISIÓN")
    aplicar_estilo_header(ws, ws.cell(row=row, column=1), "70AD47")
    ws.merge_cells(f'A{row}:G{row}')
    
    fortalezas = [
        "",
        "✓ ESTRUCTURA CLARA: Proporciona una visualización clara y sistemática del proceso de toma de decisiones",
        "✓ ANÁLISIS PROBABILÍSTICO: Permite incorporar la incertidumbre mediante probabilidades de diferentes escenarios",
        "✓ FLEXIBILIDAD: Facilita la evaluación de múltiples alternativas y escenarios simultáneamente",
        "✓ TRAZABILIDAD: Cada rama del árbol muestra claramente las consecuencias de cada decisión",
        "✓ VALOR DE LA INFORMACIÓN: Permite cuantificar el valor de obtener información adicional",
        "✓ COMUNICACIÓN EFECTIVA: Facilita la explicación y justificación de decisiones a stakeholders"
    ]
    
    for fortaleza in fortalezas:
        row += 1
        ws.cell(row=row, column=1, value=fortaleza)
        if fortaleza.startswith("✓"):
            ws.cell(row=row, column=1).font = Font(color="70AD47")
            ws.merge_cells(f'A{row}:G{row}')
    
    # Limitaciones del método
    row += 2
    ws.cell(row=row, column=1, value="LIMITACIONES DEL MÉTODO DE ÁRBOLES DE DECISIÓN")
    aplicar_estilo_header(ws, ws.cell(row=row, column=1), "C55A5A")
    ws.merge_cells(f'A{row}:G{row}')
    
    limitaciones = [
        "",
        "⚠ DEPENDENCIA DE PROBABILIDADES: Los resultados son muy sensibles a las probabilidades asignadas",
        "⚠ SIMPLICIDAD: Puede no capturar toda la complejidad de decisiones empresariales reales",
        "⚠ ESTÁTICA: No considera cambios en el tiempo ni retroalimentación dinámica",
        "⚠ SUBJETIVIDAD: Las probabilidades asignadas pueden ser subjetivas o imprecisas",
        "⚠ COMPLEJIDAD CRECIENTE: Se vuelve difícil de manejar con muchas variables y alternativas",
        "⚠ SUPUESTOS: Asume que los decision-makers son neutrales al riesgo"
    ]
    
    for limitacion in limitaciones:
        row += 1
        ws.cell(row=row, column=1, value=limitacion)
        if limitacion.startswith("⚠"):
            ws.cell(row=row, column=1).font = Font(color="C55A5A")
            ws.merge_cells(f'A{row}:G{row}')
    
    # Consideraciones adicionales
    row += 2
    ws.cell(row=row, column=1, value="CONSIDERACIONES ADICIONALES PARA LA IMPLEMENTACIÓN")
    aplicar_estilo_header(ws, ws.cell(row=row, column=1), "9900FF")
    ws.merge_cells(f'A{row}:G{row}')
    
    consideraciones = [
        "",
        "• ANÁLISIS DE SENSIBILIDAD: Realizar análisis de sensibilidad variando las probabilidades ±10%",
        "• FACTORES CUALITATIVOS: Considerar aspectos como infraestructura, competencia local, regulaciones",
        "• REVISIÓN PERIÓDICA: Establecer un cronograma para revisar y actualizar las proyecciones",
        "• PLAN DE CONTINGENCIA: Desarrollar planes alternativos para escenarios de baja demanda",
        "• MONITOREO CONTINUO: Implementar indicadores clave para monitorear el desempeño post-implementación"
    ]
    
    for consideracion in consideraciones:
        row += 1
        ws.cell(row=row, column=1, value=consideracion)
        if consideracion.startswith("•"):
            ws.merge_cells(f'A{row}:G{row}')
    
    # Información del análisis
    row += 3
    ws.cell(row=row, column=1, value="INFORMACIÓN DEL ANÁLISIS")
    aplicar_estilo_header(ws, ws.cell(row=row, column=1), "4472C4")
    ws.merge_cells(f'A{row}:C{row}')
    
    info = [
        ("Fecha de análisis:", "2024"),
        ("Inversión inicial evaluada:", f"${INVERSION_INICIAL:,} USD"),
        ("Métodos aplicados:", "6 criterios de decisión + árbol de decisiones"),
        ("Escenarios evaluados:", "3 niveles de demanda"),
        ("Ciudades analizadas:", "Santa Marta, Barranquilla, Cartagena")
    ]
    
    for etiqueta, valor in info:
        row += 1
        ws.cell(row=row, column=1, value=etiqueta)
        ws.cell(row=row, column=2, value=valor)
        ws.cell(row=row, column=1).font = Font(bold=True)

def main():
    """Función principal"""
    print("Generando análisis de expansión para Corporación XYZ...")
    
    # Crear workbook
    wb = crear_workbook()
    
    # Guardar archivo
    filename = "/home/runner/work/intro-to-github/intro-to-github/Analisis_Expansion_XYZ_Costa_Caribe.xlsx"
    wb.save(filename)
    
    print(f"Análisis completado y guardado en: {filename}")
    print("\nArchivo generado con las siguientes hojas:")
    print("1. Datos Base - Información general y datos del problema")
    print("2. Criterios Decisión - Aplicación de 6 métodos de decisión")
    print("3. Árbol Decisiones - Análisis mediante árbol de decisiones")
    print("4. Con Estudio Técnico - Evaluación con nuevas probabilidades")
    print("5. Recomendaciones - Conclusiones y recomendaciones finales")

if __name__ == "__main__":
    main()